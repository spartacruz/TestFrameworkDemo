import softest
from openpyxl import Workbook, load_workbook
import csv


class Utilities(softest.TestCase):
    def assertListItemText(self, list, value):

        for stop in list:
            print("The text is: " + stop.text)

            try:
                self.soft_assert(self.assertEqual, stop.text, value)
            except AssertionError:
                pass

            # try:
            #     assert stop.text == value
            # except AssertionError:
            #     pass

            if stop.text == value:
                print("Test Passed")
            else:
                print("Test Failed")

    def read_data_from_excel(filename, sheet):
        datalist = []
        wb = load_workbook(filename=filename)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)

            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        datalist = []
        csvdata = open(filename, "r")
        reader = csv.reader(csvdata)
        #skip header
        next(reader)
        #Iterating each row
        for row in reader:
            datalist.append(row)
        return datalist